import cmath
import random
import pygame, math, sys, numpy
from pygame.locals import *
from math import sin, cos, pi, log, sqrt, acos, atan# log(number, base), standard base is e
from numpy import array, dot, matrix, matmul
from operator import sub, add
import matplotlib.pyplot
import openpyxl


class Button:
    def __init__(self, position, width, height, PrimaryColour, function):
        self.position = position
        self.rect = Rect(position[0], position[1], width, height)
        self.shadow1 = Rect(position[0]+2, position[1]+2, width, height)
        self.shadow2 = Rect(position[0]+4, position[1]+4, width, height)
        self.shadow3 = Rect(position[0]+6, position[1]+6, width, height)
        self.Pcolour = PrimaryColour
        self.shadow1Colour = (int(PrimaryColour[0]*0.6), int(PrimaryColour[1]*0.6), int(PrimaryColour[2]*0.6))
        self.shadow2Colour = (int(PrimaryColour[0]*0.4), int(PrimaryColour[1]*0.4), int(PrimaryColour[2]*0.4))
        self.shadow3Colour = (int(PrimaryColour[0] * 0.2), int(PrimaryColour[1] * 0.2), int(PrimaryColour[2] * 0.2))
        self.function = function
        self.state = False
        self.timer = 0


    def draw(self):
        if self.state:
            colour = self.shadow1Colour
            self.timer = 10
        elif self.timer > 0:
            colour = self.shadow1Colour
        else:
            colour = self.Pcolour

        pygame.draw.rect(screen, self.shadow3Colour, self.shadow3, border_radius=8)
        pygame.draw.rect(screen, self.shadow2Colour, self.shadow2, border_radius=8)
        pygame.draw.rect(screen, self.shadow1Colour, self.shadow1, border_radius=8)
        pygame.draw.rect(screen, colour, self.rect, border_radius=5)
        self.timer -= 1


    def action(self,variable):
        return variable + self.function

    def check_click(self, mousePos):
        if self.rect.collidepoint(mousePos):
            self.state = True

    def set_state(self):
        self.state = False


class Slider:
    def __init__(self, position, minvalue, range, initialPos, name):# 0<initialPos<1
        self.button = pygame.Rect(position[0] + initialPos*150 - 20, position[1], 20, 15)  # coords, dimensions
        self.outline = pygame.Rect(position[0], position[1], 150, 15)
        self.border = pygame.Rect(position[0]-2, position[1]-2, 154, 19)
        self.range = range
        self.minvalue = minvalue
        self.sliding = False
        self.value = minvalue
        self.position = position
        self.name = name

    def draw(self):
        pygame.draw.rect(screen, LIGHTGREY, self.border)
        pygame.draw.rect(screen, DARKBLUE, self.outline)
        pygame.draw.rect(screen, SPACEYELLOW, self.button)
        self.calcValue()
        text = font.render((self.name + ' = ' + str(int(self.value))), True, LIGHTGREY)#draw value above slider
        screen.blit(text, self.position)

    def check_click(self, mousePos, mouseState):
        if self.button.collidepoint(mousePos) and mouseState:
            self.sliding = True
        else:
            self.sliding = False

    def move(self, mouseChange):
        if self.button.left >= self.outline.left and self.button.left + self.button.width <= self.outline.left + self.outline.width:
            self.button = self.button.move(mouseChange[0], 0)
        if self.button.left < self.outline.left:
            self.button.left = self.outline.left
        elif self.button.left + self.button.width > self.outline.left + self.outline.width:
            self.button.left = self.outline.left + self.outline.width - self.button.width

    def calcValue(self):
        self.value = self.minvalue + self.range*(self.button.left - self.outline.left) / (self.outline.width - self.button.width)


class Planet:
    def __init__(self, position, velocity, radius, density, primarycolour, secondarycolour, name, spinRate):
        self.position = position
        self.velocity = velocity# m/s
        self.radius = radius# m
        self.density = density# kg/m3
        self.mass = 4 * pi * density * radius**3 / 3
        self.visibleradius = log(radius * planetscale + 1, 2) + 4# no need to int() when in 3D
        self.pcolour = primarycolour
        self.scolour = secondarycolour
        self.spinRate = spinRate
        self.points = []
        self.faces = sphereFaces.copy()
        for point in spherePoints:
            self.points.append(scalarXvector(self.visibleradius, point))
        self.originalpoints = self.points.copy()
        self.scaledposition = scalarXvector(0.25 * (calcMagnitude(self.position) ** 0.25), calcUnitVector(self.position))
        self.name = name
        self.smallRenderedName = font.render(self.name, True, WHITE)
        self.renderedName = bigfont.render(name, True, SPACEYELLOW)
        self.initialDensity = density
        self.initialRadius = radius
        self.radiusSlider = Slider((12, 290), 0.1, 10*radius, 2/9, 'RADIUS')
        self.densitySlider = Slider((12, 320), 0.1, 10*density, 2/9, 'DENSITY')
        self.visitButton = Button((170, 290), 70, 50, SPACEYELLOW, None)
        self.localG = G * self.mass / radius**2
        self.faceMap = None
        self.ringPoints = None
        self.originalRingPoints = None
        self.ringWidth = None
        self.facts = []
        self.rotatedAngle = 0
        self.orbitCount = 0
        self.stepCount = 0# counts how many frames have passed for one orbit
        self.distanceSum = 0
        self.orbitTimeData = []
        self.orbitDistanceData = []
        self.custom = False
        self.redSlider = None
        self.greenSlider = None
        self.blueSlider = None
        self.deleteButton = None

    def set_facts(self, facts):
        self.facts.extend(facts)

    def display_facts(self):
        row = 0
        for fact in self.facts:
            factText = font.render(fact, True, DARKBLUE)
            screen.blit(factText, (16, 400 + row*15))
            row += 1

    def spin(self):
        rotationMatrix = array(
            [[cos(self.spinRate), -sin(self.spinRate), 0],
             [sin(self.spinRate), cos(self.spinRate), 0],
             [0, 0, 1]]
        )

        oldPoints = self.originalpoints.copy()
        self.originalpoints = []
        for point in oldPoints:
            self.originalpoints.append(matmul(rotationMatrix, point))

    def rotateAndScale(self, angleX, angleY):
        rotatedPosition = self.rotateX(self.position, angleX)
        rotatedPosition = self.rotateY(rotatedPosition, angleY)
        self.scaledposition = scalarXvector(distancescale * (calcMagnitude(rotatedPosition) ** 0.25),
                                       calcUnitVector(rotatedPosition))# scale that is non-linear but looks good

    def draw(self):
        coords = []
        for point in self.points:
            newpoint = tuple(map(add, point, self.scaledposition))
            coords.append(get2Dpoint(newpoint))

        for face in self.faces:
            createFaceShaded((coords[face[0]], coords[face[1]], coords[face[2]]),
                             self.pcolour, self.scaledposition)

    def accelerate(self):
        self.velocity = tuple(map(add,  calcAcceleration(sun, self),  self.velocity))

    def move(self):
        # movement of planet
        self.position = tuple(map(add, self.velocity, self.position))

        # find data about orbit
        oldAngle = self.rotatedAngle
        polarPosition = cmath.polar(complex(self.position[0], self.position[1]))
        self.rotatedAngle = polarPosition[1]
        self.distanceSum += polarPosition[0]
        self.stepCount += 1
        if self.rotatedAngle < 0:
            self.rotatedAngle += 2*pi
        if oldAngle - self.rotatedAngle > 6:
            self.orbitCount += 1
            if self.orbitCount > 0:# custom planets start on -1 orbits
                avgDistance = self.distanceSum / self.stepCount
                self.orbitTimeData.append((timescale * self.stepCount / (60*60*24*365.25))**2)# years^2
                self.orbitDistanceData.append((avgDistance / 1000)**3)# kilometres^3
            self.stepCount = 0
            self.distanceSum = 0

    def rotateX(self, point, angle):
        centrePoint = array(point)
        rotationmatrix = array(
            [[1, 0, 0],
             [0, cos(angle), -sin(angle)],
             [0, sin(angle), cos(angle)]])
        return matmul(rotationmatrix, centrePoint)

    def rotateY(self, point, angle):
       centrePoint = array(point)
       rotationmatrix = array(
           [[cos(angle), 0, -sin(angle)],
            [0, 1, 0],
            [sin(angle), 0, cos(angle)]])
       return matmul(rotationmatrix, centrePoint)

    def set_visibleradius(self):
        self.visibleradius = int(log(self.radius * planetscale + 1, 2)) + 4
        self.points = []
        for point in spherePoints:
            self.points.append(scalarXvector(self.visibleradius, point))

    def rotateModel(self, xangle, yangle):
        Xmatrix = array(
            [[1, 0, 0],
             [0, cos(xangle), -sin(xangle)],
             [0, sin(xangle), cos(xangle)]])
        Ymatrix = array(
           [[cos(yangle), 0, -sin(yangle)],
            [0, 1, 0],
            [sin(yangle), 0, cos(yangle)]])

        self.points = []
        for point in self.originalpoints:
            newpoint = matmul(Xmatrix, point)
            self.points.append(matmul(Ymatrix, newpoint))

    def set_radius(self, newRadius):
        oldRadius = self.visibleradius
        self.visibleradius = int(log(self.radius * planetscale + 1, 2)) + 4
        ratio = self.visibleradius / oldRadius
        self.radius = newRadius
        oldPoints = self.originalpoints.copy()
        self.originalpoints = []
        for point in oldPoints:
            self.originalpoints.append(scalarXvector(ratio, point))
        self.mass = 4 * pi * self.density * self.radius**3 / 3

    def set_density(self, newDensity):
        self.density = newDensity
        self.mass = 4 * pi * newDensity * self.radius**3 / 3

    def visit(self):
        visiting = True
        exitText = bigfont.render('CLICK ANYWHERE TO RETURN TO SIMULATION', True, WHITE)
        surface1 = pygame.Rect(0, height * 0.75, width + menuwidth, height * 0.25)
        surface2 = pygame.Rect(0, height * 0.8, width + menuwidth, height * 0.2)
        image = pygame.image.load('Astronaut.png').convert_alpha()
        scale = 60
        playerWidth = int(0.5 * scale)
        playerHeight = 2 * scale
        player = pygame.Rect(width/2 - playerWidth/2, height*0.75 - playerHeight - 400, playerWidth, playerHeight)
        g = G * self.mass / (self.radius**2) * scale # g measured in pixels
        forceVR = 0
        mass = 60
        velocity = -2
        while visiting:
            framerate = clock.get_fps()
            keysPressed = pygame.key.get_pressed()
            forceVR = -g * mass
            if keysPressed[K_RIGHT]:
                player.x += 4
            elif keysPressed[K_LEFT]:
                player.x -= 4
            for event in pygame.event.get():

                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

                if event.type == pygame.MOUSEBUTTONDOWN:
                    visiting = False
                if event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_SPACE:
                        velocity = 3.16228
            velocity += -g / framerate**2
            player.y -= velocity
            if player.y >= height*0.75 - playerHeight:
                player.y = height*0.75 - playerHeight
                velocity = 0
            if random.randint(1, 150) == 1:
                stars.pop(0)
                stars.append((random.randint(menuwidth, width + menuwidth), random.randint(0, height)))
            screen.fill(BLACK)
            for star in stars:
                pygame.draw.circle(screen, WHITE, star, 1)
            pygame.draw.rect(screen, self.scolour, surface1)
            pygame.draw.rect(screen, self.pcolour, surface2)
            screen.blit(image, player)
            heightText = font.render('height = ' + str(round((height*0.75 - player.y - playerHeight) / scale,
                                                             1)) + ' metres', True, WHITE)
            if player.y > 0:
                screen.blit(heightText, (player.x+30, player.y+10))
            else:
                screen.blit(heightText, (player.x+30, 0))
            screen.blit(exitText, (50, 50))
            pygame.display.update()


class RingedPlanet(Planet):

    def draw(self):
        coords = []
        for point in self.points:
            newpoint = tuple(map(add, point, self.scaledposition))
            coords.append(get2Dpoint(newpoint))

        listLength = len(self.ringPoints)
        lines = []
        for i in range(listLength):
            line = (self.ringPoints[i], self.ringPoints[((i + 1) % listLength)])
            lines.append(line)
        frontLines = []
        backLines = []
        for line in lines:
            if get_midpoint(line[0], line[1])[2] > 0:
                backLines.append(line)
            else:
                frontLines.append(line)

        for line in backLines: # draw back half of ring that could be obscured by planet
            point1 = get2Dpoint(tuple(map(add, line[0], self.scaledposition)))
            point2 = get2Dpoint(tuple(map(add, line[1], self.scaledposition)))
            pygame.draw.line(screen, self.scolour, (point1[0], point1[1]), (point2[0], point2[1]), self.ringWidth)

        for face in self.faces: # draw planet
            createFaceShaded((coords[face[0]], coords[face[1]], coords[face[2]]),
                             self.pcolour, self.scaledposition)

        for line in frontLines: # draw front half of ring
            point1 = get2Dpoint(tuple(map(add, line[0], self.scaledposition)))
            point2 = get2Dpoint(tuple(map(add, line[1], self.scaledposition)))
            pygame.draw.line(screen, self.scolour, (point1[0], point1[1]), (point2[0], point2[1]), self.ringWidth)

    def set_rings(self, width):
        self.ringPoints = []
        for q in range(7, 13):
            self.ringPoints.append(scalarXvector(1.4, self.originalpoints[q]))
        self.originalRingPoints = self.ringPoints.copy()
        self.ringWidth = width

    def rotateModel(self, xangle, yangle):
        Xmatrix = array(
            [[1, 0, 0],
             [0, cos(xangle), -sin(xangle)],
             [0, sin(xangle), cos(xangle)]])
        Ymatrix = array(
            [[cos(yangle), 0, -sin(yangle)],
             [0, 1, 0],
             [sin(yangle), 0, cos(yangle)]])

        self.points = []
        for point in self.originalpoints:
            newpoint = matmul(Xmatrix, point)
            self.points.append(matmul(Ymatrix, newpoint))
        self.ringPoints = []
        for point in self.originalRingPoints:
            newpoint = matmul(Xmatrix, point)
            self.ringPoints.append(matmul(Ymatrix, newpoint))

    def spin(self):
        rotationMatrix = array(
            [[cos(self.spinRate), -sin(self.spinRate), 0],
             [sin(self.spinRate), cos(self.spinRate), 0],
             [0, 0, 1]]
        )

        oldPoints = self.originalpoints.copy()
        self.originalpoints = []
        for point in oldPoints:
            self.originalpoints.append(matmul(rotationMatrix, point))
        oldPoints = self.originalRingPoints.copy()
        self.originalRingPoints = []
        for point in oldPoints:
            self.originalRingPoints.append(matmul(rotationMatrix, point))


class MappedPlanet(Planet):
    def draw(self):
        coords = []
        for point in self.points:
            newpoint = tuple(map(add, point, self.scaledposition))
            coords.append(get2Dpoint(newpoint))

        for i in range(len(self.faces)):
            if self.faceMap[i]:
                colour = self.scolour
            else:
                colour = self.pcolour
            createFaceShaded((coords[self.faces[i][0]], coords[self.faces[i][1]], coords[self.faces[i][2]]),
                             colour, self.scaledposition)


class Star(Planet):
    def draw(self):
        endPointUp = tuple(map(add, self.scaledposition, scalarXvector(2, self.points[0])))
        endPointDown = tuple(map(add, self.scaledposition, scalarXvector(2, self.points[-1])))
        lineEndPointUp = get2Dpoint(endPointUp)
        lineEndPointDown = get2Dpoint(endPointDown)
        lineStartPointUp = get2Dpoint(self.points[0])
        lineStartPointDown = get2Dpoint(self.points[-1])
        order = True
        if tuple(map(sub, endPointUp, endPointDown))[2] < 0:
            order = False

        if order and arrows:
            pygame.draw.line(screen, GREEN, (lineEndPointUp[0], lineEndPointUp[1]),
                             (lineStartPointUp[0], lineStartPointUp[1]), 2)
        elif arrows:
            pygame.draw.line(screen, GREEN, (lineEndPointDown[0], lineEndPointDown[1]),
                             (lineStartPointDown[0], lineStartPointDown[1]), 2)

        coords = []
        for point in self.points:
            newpoint = tuple(map(add, point, self.scaledposition))
            coords.append(get2Dpoint(newpoint))
        for face in self.faces:
            createFaceOutlined((coords[face[0]], coords[face[1]], coords[face[2]]),
                               self.pcolour, self.scolour)

        if order and arrows:
            pygame.draw.line(screen, GREEN, (lineEndPointDown[0], lineEndPointDown[1]),
                             (lineStartPointDown[0], lineStartPointDown[1]), 2)
        elif arrows:
            pygame.draw.line(screen, GREEN, (lineEndPointUp[0], lineEndPointUp[1]),
                             (lineStartPointUp[0], lineStartPointUp[1]), 2)


class Meteor:
    def __init__(self, position, velocity):
        self.position = position
        self.velocity = scalarXvector(timescale, velocity)
        self.mass = 1000
        self.scaledposition = (0, 0, 0)
        self.resultantAcceleration = (0, 0, 0)
        self.rotatedAccelerationVector = (0, 0, 0)

    def accelerate(self, planets):
        accelerationVector = (0, 0, 0)
        for planet in planets:
            accelerationVector = tuple(map(add, accelerationVector, calcAcceleration(planet, self)))
        self.velocity = tuple(map(add, accelerationVector, self.velocity))
        self.resultantAcceleration = accelerationVector

    def move(self):
        self.position = tuple(map(add, self.position, self.velocity))

    def draw(self):
        scaledPosition = get2Dpoint(self.scaledposition)
        endPoint = get2Dpoint(tuple(map(add, self.scaledposition,
                                        scalarXvector(18, calcUnitVector(self.rotatedAccelerationVector)))))
        if arrows:
            if self.rotatedAccelerationVector[2] > 0:#depth effect
                pygame.draw.line(screen, PURPLE, (scaledPosition[0], scaledPosition[1]), (endPoint[0], endPoint[1]), 1)
                pygame.draw.circle(screen, ORANGE, (scaledPosition[0], scaledPosition[1]), 3)
            else:
                pygame.draw.circle(screen, ORANGE, (scaledPosition[0], scaledPosition[1]), 3)
                pygame.draw.line(screen, PURPLE, (scaledPosition[0], scaledPosition[1]), (endPoint[0], endPoint[1]), 1)
        else:
            pygame.draw.circle(screen, ORANGE, (scaledPosition[0], scaledPosition[1]), 3)

    def rotateAndScale(self, angleX, angleY):
        rotatedPosition = self.rotateX(self.position, angleX)
        rotatedPosition = self.rotateY(rotatedPosition, angleY)
        self.scaledposition = scalarXvector(distancescale * (calcMagnitude(rotatedPosition) ** 0.25),
                                       calcUnitVector(rotatedPosition))  # scale that is non-linear but looks good
        rotatedVector = self.rotateX(self.resultantAcceleration, angleX)
        self.rotatedAccelerationVector = self.rotateY(rotatedVector, angleY)

    def rotateX(self, point, angle):
        centrePoint = array(point)
        rotationmatrix = array(
            [[1, 0, 0],
             [0, cos(angle), -sin(angle)],
             [0, sin(angle), cos(angle)]])
        return matmul(rotationmatrix, centrePoint)

    def rotateY(self, point, angle):
        centrePoint = array(point)
        rotationmatrix = array(
            [[cos(angle), 0, -sin(angle)],
             [0, 1, 0],
             [sin(angle), 0, cos(angle)]])
        return matmul(rotationmatrix, centrePoint)


def get2Dpoint(point):
    (x, y, z) = tuple(map(add, (point[0], point[1], point[2]), (0, 4, 1000)))
    projectedX = menuwidth + width / 2 + ((x * distance) / (z + distance)) * scale
    projectedY = height / 2 + ((y * distance) / (z + distance)) * scale
    try: return int(projectedX), int(projectedY), x, y, z
    except ValueError:
        return int(point[0]+width/2), int(point[1]+height/2), x, y, z


def createFaceOutlined(coordinates, colour, outlineColour):
    a, b, c = coordinates[0], coordinates[1], coordinates[2]  # points of the triangle
    limit = distance / scale
    if a[4] < limit:  # if the face is too close (distance/scale)
        return
    elif b[4] < limit:
        return
    elif c[4] < limit:
        return
    vector1 = (b[2] - a[2], b[3] - a[3], b[4] - a[4],)  # uses the 3D coords(not the 2D projected point)

    vector2 = (c[2] - a[2], c[3] - a[3], c[4] - a[4],)  # uses the 3D coords(not the 2D projected point)
    crossVector = crossProduct(vector1, vector2)

    if crossVector[2] < 0:  # calculate if the z part vector of the face points away from the screen
        return

    coords = ((a[0], a[1]), (b[0], b[1]), (c[0], c[1]),)  # the first two items in the tuple are the 2D coords
    pygame.draw.polygon(screen, colour, coords)  # draw face
    pygame.draw.polygon(screen, outlineColour, coords, 1)  # draw outline


def createFaceShaded(coordinates, colour, planetPosition):
    a, b, c = coordinates[0], coordinates[1], coordinates[2]  # points of the triangle
    limit = distance / scale
    if a[4] < limit:  # if the face is too close (distance/scale)
        return
    elif b[4] < limit:
        return
    elif c[4] < limit:
        return
    vector1 = (b[2] - a[2], b[3] - a[3], b[4] - a[4],)  # uses the 3D coords(not the 2D projected point)

    vector2 = (c[2] - a[2], c[3] - a[3], c[4] - a[4],)  # uses the 3D coords(not the 2D projected point)
    crossVector = crossProduct(vector1, vector2)

    if crossVector[2] < 0:  # calculate if the z part vector of the face points away from the screen
        return

    colour = shade(planetPosition, crossVector, colour)

    coords = ((a[0], a[1]), (b[0], b[1]), (c[0], c[1]),)  # the first two items in the tuple are the 2D coords
    pygame.draw.polygon(screen, colour, coords)  # draw face


def crossProduct(vector1, vector2):#3D cross product
    return [vector1[1]*vector2[2] - vector1[2]*vector2[1],
            vector1[2]*vector2[0] - vector1[0]*vector2[2],
            vector1[0]*vector2[1] - vector1[1]*vector2[0]
            ]


def dotProduct(vector1, vector2):
    return vector1[0]*vector2[0] + vector1[1]*vector2[1] + vector1[2]*vector2[2]


def calcAcceleration(star, planet):
    global timescale
    distanceVector = tuple(map(sub, star.position, planet.position))
    forceMagnitude = (G * star.mass * planet.mass) / (calcMagnitude(distanceVector)**2)
    return scalarXvector(timescale**2 * forceMagnitude/planet.mass, calcUnitVector(distanceVector))


def calcMagnitude(vector):
    return  sqrt(vector[0]**2 + vector[1]**2 + vector[2]**2)


def calcUnitVector(vector):
    magnitude = calcMagnitude(vector)
    if magnitude != 0:
        return vector[0]/magnitude , vector[1]/magnitude, vector[2]/magnitude
    else: return 0,0,0


def scalarXvector(scalar, vector):
    return  scalar*vector[0], scalar*vector[1], scalar*vector[2]


def shade(position, normal, colour):
    angle = acos(dotProduct(position, normal) / (calcMagnitude(position)*calcMagnitude(normal)))
    return scalarXvector(1-(angle / pi), colour)


def get_midpoint(point1, point2):
    return scalarXvector(0.5, tuple(map(add, point1, point2)))


def create_meteor(list):
    velocityX = random.randrange(-2 * 10**4, 2 * 10**4)
    velocityY = random.randrange(-2 * 10**4, 2 * 10**4)
    velocityZ = random.randrange(-2 * 10**4, 2 * 10**4)

    positionX = random.randrange(-10**12, 10**12)
    positionY = random.randrange(-10**12, 10**12)
    positionZ = random.randrange(-10**12, 10**12)

    newmeteor = Meteor((positionX, positionY, positionZ), (velocityX, velocityY, velocityZ))
    list.append(newmeteor)


def average(list):
    return sum(list) / len(list)


def plot(star, log):
    xdata = []
    ydata = []
    pointColours = []
    text = 'linear scale'
    matplotlib.pyplot.style.use('dark_background')
    if log:
        matplotlib.pyplot.xscale('log')
        matplotlib.pyplot.yscale('log')
        text = 'logarithmic scale'

    for planet in planets:
        if len(planet.orbitTimeData) > 0:
            x = average(planet.orbitTimeData)
            y = average(planet.orbitDistanceData) / 10 ** 6
            xdata.append(x)
            ydata.append(y)
            pointColours.append(scalarXvector(1 / 255, planet.pcolour))  # matplotlib uses rgb values in range 0-1
            matplotlib.pyplot.text(x, y, planet.name)

    # matplotlib.pyplot.plot(xdata, ydata, color=(1, 1, 0.9))# line connecting points (looks like a straight line)
    matplotlib.pyplot.plot([0, max(xdata)], [0,  max(xdata) * G * star.mass / (4 * pi**2)], color=(1, 1, 0.8))
    matplotlib.pyplot.scatter(x=xdata, y=ydata, c=pointColours)# points
    matplotlib.pyplot.xlabel('Orbital Time Period squared (Earth years^2)')
    matplotlib.pyplot.ylabel('Mean orbital distance cubed(10^6 km^3)')
    matplotlib.pyplot.title('Orbital data(' + text + ')')
    matplotlib.pyplot.show()


def create_planet(star):
    global planetNumber, currentPlanet, sliders
    name = ''
    typing = True
    nameText = bigfont.render('NAME YOUR PLANET: ' + name, True, LIGHTGREY)
    screen.blit(nameText, (266, 216))
    pygame.display.update((260, 210, 0, 50))
    while typing:
        for event in pygame.event.get():
            if event.type == pygame.KEYDOWN:
                if event.key == K_RETURN:
                    typing = False
                elif event.key == K_BACKSPACE:
                    name = name[:-1]
                    nameText = bigfont.render('NAME YOUR PLANET: ' + name, True, LIGHTGREY)
                else:
                    name += event.unicode
                    nameText = bigfont.render('NAME YOUR PLANET: ' + name, True, LIGHTGREY)
        screen.blit(menuImage, (0, 0))
        screen.blit(title, (6, 14))
        pygame.draw.rect(screen, SPACEYELLOW, menuborder, width=borderwidth, border_radius=2 * borderwidth)
        pygame.draw.rect(screen, LIGHTGREY, factBox, border_radius=10)
        pygame.draw.rect(screen, DARKBLUE, (260, 210, 700, 50))
        pygame.draw.rect(screen, SPACEYELLOW, (260, 210, 700, 50), 5, 3)
        screen.blit(nameText, (266, 216))
        pygame.display.update((260, 210, 0, 50))

    boolean = [True, False]
    Pcolour = (random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
    Scolour = (random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
    density = random.randint(500, 1500)
    radiusExponent = 4 + (random.random() * 7.5)
    radius = 10**radiusExponent
    angle = 2 * pi * random.random()
    rotationMatrix = array([[cos(angle), -sin(angle), 0],
                            [sin(angle), cos(angle), 0],
                            [0, 0, 1]
                            ])
    orbitalRadiusExponent = 11 + 1.5 * random.random()
    orbitalRadius = 0.4 * 10**orbitalRadiusExponent
    speed = sqrt(G * star.mass / orbitalRadius)
    position = matmul(rotationMatrix, (orbitalRadius, 0, 0))
    velocity = matmul(rotationMatrix, (0, speed * timescale, 0))
    spinRate = 0.1 * (random.random() - 0.5)
    if random.choice(boolean):
        newPlanet = RingedPlanet(position, velocity, radius, density, Pcolour, Scolour, name, spinRate)
        newPlanet.set_rings(random.randint(1, 5))
    else:
        newPlanet = Planet(position, velocity, radius, density, Pcolour, Scolour, name, spinRate)
    planets.append(newPlanet)
    planetsToDraw.append(newPlanet)
    selectablePlanets.append(newPlanet)
    newPlanet.custom = True
    newPlanet.orbitCount = -1
    # ^ the planet must pass through start line before data is collected to allow accurate time measurement
    newPlanet.set_facts(['CUSTOM PLANET'])
    planetNumber = len(selectablePlanets) - 1
    currentPlanet = selectablePlanets[planetNumber]
    newPlanet.redSlider = Slider((20, 440), 0, 255, (Pcolour[0] / 255), 'RED')
    newPlanet.greenSlider = Slider((20, 460), 0, 255, Pcolour[1] / 255, 'GREEN')
    newPlanet.blueSlider = Slider((20, 480), 0, 255, Pcolour[2] / 255, 'BLUE')
    sliders = [newPlanet.radiusSlider, newPlanet.densitySlider,
               newPlanet.redSlider, newPlanet.greenSlider, newPlanet.blueSlider]
    newPlanet.deleteButton = Button((164, 610), 60, 30, RED2, None)


def spreadsheet(planets):
    wb = openpyxl.Workbook()
    ws = wb.active
    '''for row in ws:
        for cell in row:
            cell.value = None# clear worksheet'''

    ws.cell(row=1, column=1, value='Planet Name')
    ws.cell(row=1, column=2, value='Average orbital time period (days)')
    ws.cell(row=1, column=3, value='Average orbital distance (km)')
    ws.cell(row=1, column=4, value='Mass (kg)')
    ws.cell(row=1, column=5, value='Radius (km)')
    ws.cell(row=1, column=6, value='Density (kg m^-3)')
    ws.cell(row=1, column=7, value='Surface gravity (m s^-2)')

    ws.cell(row=2, column=1, value=sun.name)
    ws.cell(row=2, column=2, value='N/A')
    ws.cell(row=2, column=3, value='N/A')
    ws.cell(row=2, column=4, value=sun.mass)
    ws.cell(row=2, column=5, value=sun.radius * 10 ** -3)
    ws.cell(row=2, column=6, value=sun.density)
    ws.cell(row=2, column=7, value=G * sun.mass / (sun.radius ** 2))

    count = 3
    for planet in planets:
        ws.cell(row=count, column=1, value=planet.name)
        if len(planet.orbitTimeData) > 0:
            ws.cell(row=count, column=2, value=average(planet.orbitTimeData)**0.5 * 365.25)
            ws.cell(row=count, column=3, value=average(planet.orbitDistanceData)**(1/3) / 10 ** 6)
        ws.cell(row=count, column=4, value=planet.mass)
        ws.cell(row=count, column=5, value=planet.radius * 10**-3)
        ws.cell(row=count, column=6, value=planet.density)
        ws.cell(row=count, column=7, value=G * planet.mass / (planet.radius**2))
        count += 1
    wb.save('OrbitData.xlsx')

# planet info https://nssdc.gsfc.nasa.gov/planetary/factsheet/
pygame.init()
clock = pygame.time.Clock()
pygame.display.set_caption("SOLAR SYSTEM")
width = 750
height = 750
menuwidth = 250
size = width + menuwidth, height
screen = pygame.display.set_mode(size)
icon = pygame.image.load('SolarSystem.png')
menuImage = pygame.image.load('Backgrounds/doodad-2.png').convert()
pygame.display.set_icon(icon)
font = pygame.font.SysFont('Comic Sans MS', 20)
bigfont = pygame.font.SysFont('Comic Sans MS', 40)


borderwidth = 3
menubackground = menuImage.get_rect().update(0, 0, menuwidth, height)
menuborder = Rect(0, 0, menuwidth, height)
factBox = Rect(12, 350, menuwidth-24, 300)

planetscale = 4.721393815727515e-06
distancescale = 0.3
timescale = 10 **4.5
halfpi = pi/2
e = math.e
G = 6.6743 * 10 ** -11

GREY = (175, 175, 170)
RED = (255, 0, 0)
WHITE = (255, 255, 255)
BLUE = (0, 0, 255)
GREEN = (0, 255, 0)
PURPLE = (255, 0, 255)
BLACK = (0, 0, 0)
YELLOW = (255, 195, 0)
ORANGE = (255, 145, 0)
BROWN = (200, 170, 100)
PINK = (255, 153, 255)
DARKBLUE = (14, 21, 61)
SPACEYELLOW = (252, 163, 11)
LIGHTGREY = (229, 229, 229)
DARKGREY = (60, 60, 60)
menuColour = (3, 87, 92)
menuBorderColour = (202, 202, 202)
RED2 = (179, 46, 68)
arrowColour = (84, 84, 84)

factText = bigfont.render('Fun Facts:', True, BLACK)
title = bigfont.render('The Solar System', True, LIGHTGREY)
SolarSystemFacts = [font.render('The solar system has', True, DARKBLUE),
                    font.render('8 planets.', True, DARKBLUE),
                    font.render('', False, RED),
                    font.render('Pluto is not a planet.', True, DARKBLUE),
                    font.render('', False, RED),
                    font.render('The solar system is over', True, DARKBLUE),
                    font.render('4.6 billion years old.', True, DARKBLUE),
                    font.render('', False, RED),
                    font.render('The solar system is in the', True, DARKBLUE),
                    font.render('Milky Way galaxy.', True, DARKBLUE)]


boxFaces = [(0, 1, 2),(0, 2, 3),(2, 7, 3),(2, 6, 7),(5, 2, 1),(2, 5, 6),
            (4, 1, 0),(1, 4, 5),(6, 5, 4),(7, 6, 4),(3, 7, 4),(0, 3, 4),]# links to points in boxPoints
boxPoints = [(-1, -1, -1),(1, -1, -1),(1, 1, -1),(-1, 1, -1),
            (-1, -1, 1),(1, -1, 1),(1, 1, 1),(-1, 1, 1),]
tempSpherePoints = [(0, 1, 0),# N pole
                (-sqrt(3)/2, 0.5, 0), (-sqrt(3)/4, 0.5, 0.75), (sqrt(3)/4, 0.5, 0.75),# 30deg N
                (sqrt(3)/2, 0.5, 0), (sqrt(3)/4, 0.5, -0.75), (-sqrt(3)/4, 0.5, -0.75),
                (-sqrt(3)/2, 0, 0.5), (0,0,1), (sqrt(3)/2, 0, 0.5),# equator
                (sqrt(3)/2, 0, -0.5), (0,0,-1), (-sqrt(3)/2, 0, -0.5),
                (-sqrt(3)/2, -0.5, 0), (-sqrt(3)/4, -0.5, 0.75), (sqrt(3)/4, -0.5, 0.75),# 30deg S
                (sqrt(3)/2, -0.5, 0), (sqrt(3)/4, -0.5, -0.75), (-sqrt(3)/4, -0.5, -0.75),
                (0, -1, 0)# S pole
                ]
sphereFaces = [(0, 2, 1), (0, 3, 2), (0, 4, 3), (0, 5, 4), (0, 6, 5), (0, 1, 6),# 0-5
               (1, 7, 12), (1, 2, 7), (2, 8, 7), (2, 3, 8), (3, 9, 8), (3, 4, 9),# 6-11
               (4, 10, 9), (4, 5, 10), (5, 11, 10), (5, 6, 11), (6, 12, 11), (1, 12, 6),# 12-17
               (13, 12, 7), (14, 13, 7), (14, 7, 8), (15, 14, 8), (15, 8, 9), (16, 15, 9),# 18-23
               (16, 9, 10), (17, 16, 10), (17, 10, 11), (18, 17, 11), (18, 11, 12), (13, 18, 12),# 24-29
               (19, 13, 14), (19, 14, 15), (19, 15, 16), (19, 16, 17), (19, 17, 18), (19, 18, 13)]# 30-35

spherePoints = []
X90Matrix = array([[1, 0, 0], [0, 0, -1], [0, 1, 0]])
for point in tempSpherePoints:
    spherePoints.append(matmul(X90Matrix, point))


distance = 10
scale = 80


# position, velocity, radius, density, primarycolour, secondarycolour, name, spin rate
# all measurements in SI units
mercury = Planet((0.579 * 10 ** 11, 0, 0), (0,4.74 * 10**4 * timescale, 0), (2.4345 * 10 ** 6), 5429,
                 GREY, YELLOW, 'MERCURY', 0.01)
venus = Planet((1.082 * 10 ** 11, 0, 0), (0, 3.5 * 10**4 * timescale, 0), (6.052 * 10**6), 5243,
               YELLOW, GREY, 'VENUS', -0.01)
earth = MappedPlanet((1.496 * 10 ** 11, 0,0), (0, 2.98 * 10**4 * timescale, 0), (6.378 * 10**6), 5514,
               BLUE, GREEN, 'EARTH', 0.01)
mars = Planet((2.06 * 10 ** 11, 0, 0), (0, 2.407 * 10**4 * timescale, 0), (3.3895 * 10**6), 3934,
              RED, ORANGE, 'MARS', 0.01)
jupiter = MappedPlanet((7.785 * 10**11, 0, 0), (0, 1.31 * 10**4 * timescale, 0), (71.492 * 10**6), 1326,
                 YELLOW, RED, 'JUPITER', 0.01)
saturn = RingedPlanet((1.432 * 10**12, 0, 0), (0, 0.97 * 10**4 * timescale, 0), (60.268 * 10**6), 687,
                BROWN, GREY, 'SATURN', 0.01)
uranus = Planet((2.867 * 10**12, 0, 0), (0, 0.68 * 10**4 * timescale, 0), (25.559 * 10**6), 1270,
                GREY, BLUE, 'URANUS', -0.01)
neptune = Planet((4.515 * 10**12, 0, 0), (0, 0.54 * 10**4 * timescale, 0), (24.764 * 10**6), 1638,
                 BLUE, GREY, 'NEPTUNE', 0.01)

sun = Star((0, 0, 0), (0, 0, 0), (6.9634 * 10 ** 8), 1410,
             YELLOW, ORANGE, 'SUN', 0.003)

earth.faceMap = ([True, True, False, False, False, False,
                 False, False, False, True, True, True,
                 True, True, False, True, True, False,
                 False, True, False, False, False, False,
                 False, True, False, True, True, True,
                 True, False, False, True, False, False])
jupiter.faceMap = ([False, False, False, False, False, False,
                    False, False, False, False, False, False,
                    False, False, False, False, False, False,
                    False, False, False, False, False, False,
                    False, True, True, False, False, False,
                    False, False, False, False, False, False])

saturn.set_rings(3)
mercury.set_facts(['Mercury is the smallest', 'planet in the solar system.', '',
                   "It is only slightly larger", "than the Earth's moon.", '',
                   'Mercury orbits the Sun at', 'nearly 47km/s.', '',
                   'Mercury has no moons.', '',
                   "In 1974, NASA's Mariner 10 was", "the first spacecraft to fly past", "Mercury."])
venus.set_facts(['Venus spins in the opposite', 'direction to most other planets.', '',
                 'Venus is the hottest planet', 'despite being further away from the', 'Sun than Mercury.', '',
                 "This is due to it's thick atmosphere."
                 ])
earth.set_facts(['Earth is the only known planet', 'in the universe to sustain life.', '',
                 'Earth is the only planet in the', 'solar system with liquid', "water on it's surface.", '',
                 'Earth is the only planet not', 'named after a Greek or Roman God.'])
mars.set_facts(['Mars is the only planet', 'that humans have sent rovers to.', '',
                'These rovers have found evidence', 'that suggests Mars once had', 'a thick atmosphere',
                'and possibly liquid water.', '',
                'Mars has two moons.', '',
                'Mars is known as the Red Planet', 'due to the iron oxide in the Martian', 'soil.'])
jupiter.set_facts(['Jupiter is the largest planet', 'in the solar system.', '',
                   "Jupiter's Giant Red Spot is a", 'storm larger than the Earth.', '',
                   'Jupiter has 80 moons.'])
saturn.set_facts(["Saturn's unique rings are", 'made from ice and rocks', '',
                  "Saturn's atmosphere is mostly", 'made from Hydrogen and', 'Helium.', '',
                  'Saturn is less dense', 'than water.'])
uranus.set_facts(['Uranus is mostly made of ice.', '',
                  "Uranus rotates on it's side.", '',
                  'Uranus also has rings like', 'Saturn, but they are much', 'harder to see.'])
neptune.set_facts(['Neptune is the furthest planet', 'from the Sun.', '',
                   'Neptune is the only planet not', 'visible to the naked eye.', '',
                   'The existence of Neptune was', 'predicted by mathematics.'])
sun.set_facts(['The Sun is a yellow dwarf star.', '',
               'The core of the Sun is 15 million', 'degrees Celsius.', ''
               'Nuclear fusion takes place', "in the Sun.", '',
               'The Sun is 4.6 billion', 'years old.'])


planets = [mercury, venus, earth, mars, jupiter, saturn, uranus, neptune]
planetsToDraw = [sun, mercury, venus, earth, mars, jupiter, saturn, uranus, neptune]
selectablePlanets = [None, sun, mercury, venus, earth, mars, jupiter, saturn, uranus, neptune]
planetNumber = 0
currentPlanet = selectablePlanets[planetNumber]

stars = []
for x in range(15):
    stars.append((random.randint(menuwidth, width+menuwidth), random.randint(0, height)))


meteors = []

sliders = []
mouseState = False
planetLeft = Button((75, 145), 40, 40, LIGHTGREY, -1)
planetRight = Button((140, 145), 40, 40, LIGHTGREY, 1)
cycleButtons = [planetLeft, planetRight]
plotButton = Button((16, 60), 60, 30, SPACEYELLOW, None)
plotText = font.render('PLOT', True, BLACK)
createButton = Button((95, 60), 60, 30, SPACEYELLOW, None)
createText1 = font.render('NEW', True, BLACK)
createText2 = font.render('PLANET', True, BLACK)
clearButton = Button((170, 60), 68, 30, SPACEYELLOW, None)
clearText1 = font.render('CLEAR', True, BLACK)
clearText2 = font.render('METEORS', True, BLACK)
deleteText = font.render('DELETE', True, BLACK)
visitText1 = font.render('VISIT', True, BLACK)
visitText2 = font.render('PLANET', True, BLACK)
excelButton = Button((16, 100), 60, 30, SPACEYELLOW, None)
excelText = font.render('EXCEL', True, BLACK)
hideButton = Button((95, 100), 60, 30, SPACEYELLOW, None)
hideText1 = font.render('HIDE', True, BLACK)
hideText2 = font.render('ARROWS', True, BLACK)
viewButton = Button((170, 100), 68, 30, SPACEYELLOW, None)
viewText1 = font.render('TOPDOWN', True, BLACK)
viewText2 = font.render('VIEW', True, BLACK)
buttons = [planetLeft, planetRight, plotButton, createButton, clearButton, excelButton, hideButton, viewButton]

logarithmic = False # used to alternate the scale used by the graph

bottomText1 = font.render('Press SPACE to add more', True, SPACEYELLOW)
bottomText2 = font.render('Use ARROW keys to rotate view', True, SPACEYELLOW)
bottomText3 = font.render('Polar axis', True, GREEN)
bottomText4 = font.render('Force and acceleration due to gravity', True, PURPLE)

# pygame
totalxrotation = 0
totalyrotation = 0

meteorNumber = 0
timeElapsed = 0
exitLoop = False
arrows = True

while not exitLoop:
    clock.tick(100)
    timeElapsed += timescale
    xanglechange = 0
    yanglechange = 0
    mousePos = pygame.mouse.get_pos()
    mouseChange = pygame.mouse.get_rel()
    # buttons
    for button in buttons:
        button.set_state()
    if currentPlanet is not None:
        currentPlanet.visitButton.set_state()
        if currentPlanet.custom:
            currentPlanet.deleteButton.set_state()

    for event in pygame.event.get():

        if event.type == pygame.QUIT:
            exitLoop = True

        if event.type == pygame.MOUSEBUTTONDOWN:
            #create_planet()
            for slider in sliders:
                slider.check_click(mousePos, True)
            # buttons
            for button in buttons:
                button.check_click(mousePos)
            if currentPlanet is not None:
                if currentPlanet.custom:
                    currentPlanet.deleteButton.check_click(mousePos)
                currentPlanet.visitButton.check_click(mousePos)
                if currentPlanet.visitButton.state:
                    currentPlanet.visit()

        elif event.type == pygame.MOUSEBUTTONUP:
            for slider in sliders:
                slider.check_click(mousePos, False)

        if event.type == pygame.KEYDOWN:
            if event.key == K_SPACE:
                create_meteor(meteors)
                meteorNumber += 1

    keysPressed = pygame.key.get_pressed()
    if keysPressed[K_RIGHT]:
        totalyrotation += 0.015
    elif keysPressed[K_LEFT]:
        totalyrotation -= 0.015
    if keysPressed[K_UP]:
        totalxrotation -= 0.015
    elif keysPressed[K_DOWN]:
        totalxrotation += 0.015
    if keysPressed[K_z] and scale > 1:
        scale -= 1
    elif keysPressed[K_x]:
        scale += 1

    screen.fill(BLACK)
    if random.randint(1, 150) == 1:
        stars.pop(0)
        stars.append((random.randint(menuwidth, width+menuwidth), random.randint(0, height)))
    for star in stars:
        pygame.draw.circle(screen, WHITE, star, 1)
    # drawing planets
    for planet in planets:
        planet.accelerate()
        planet.move()
        planet.rotateAndScale(totalxrotation, totalyrotation)

    planetsToDraw.sort(key=lambda planet : calcMagnitude(tuple(map(add, planet.scaledposition, (0, 4, 1000)))),
                       reverse=True)
    for meteor in meteors:
        meteor.accelerate(planetsToDraw)
        meteor.move()
        meteorPosition = meteor.position
        meteor.rotateAndScale(totalxrotation, totalyrotation)
        meteor.draw()
        if abs(meteorPosition[0]) + abs(meteorPosition[1]) + abs(meteorPosition[2]) > 10**13:
            meteors.remove(meteor)
            del meteor

    for planet in planetsToDraw:
        planet.spin()
        planet.rotateModel(totalxrotation, totalyrotation)
        planet.draw()

    for button in cycleButtons:
        if button.state:
            planetNumber += button.function
            if planetNumber > len(selectablePlanets)-1:
                planetNumber = 0
            if planetNumber < 0:
                planetNumber = len(selectablePlanets)-1
            currentPlanet = selectablePlanets[planetNumber]
            sliders = []
            if currentPlanet is not None:
                if currentPlanet.custom:
                    sliders = [currentPlanet.radiusSlider, currentPlanet.densitySlider,
                               currentPlanet.redSlider, currentPlanet.greenSlider, currentPlanet.blueSlider]
                else:
                    sliders = [currentPlanet.radiusSlider, currentPlanet.densitySlider]

    if plotButton.state and (timeElapsed / (60*60*24*365.25) > 1):
        logarithmic = not(logarithmic)
        plot(sun, logarithmic)
    if createButton.state:
        create_planet(sun)
    if clearButton.state:
        meteors = []
    if excelButton.state:
        spreadsheet(planets)
    if hideButton.state:
        arrows = not arrows
    if viewButton.state:
        totalxrotation = 0
        totalyrotation = 0

    if currentPlanet is not None:# physics side
        planetPosition = get2Dpoint(currentPlanet.scaledposition)

        # line to show direction of force of g
        startPoint = tuple(map(add, currentPlanet.scaledposition,# start is on planets surface (not planet centre)
                               scalarXvector(-currentPlanet.visibleradius,
                                             calcUnitVector(currentPlanet.scaledposition))))
        endPoint = tuple(map(add, currentPlanet.scaledposition,
                             scalarXvector(-50, calcUnitVector(currentPlanet.scaledposition))))
        lineStartPoint = get2Dpoint(startPoint)
        lineEndPoint = get2Dpoint(endPoint)

        if arrows:# draw lines
            pygame.draw.line(screen, PURPLE, (lineEndPoint[0], lineEndPoint[1]),
                             (lineStartPoint[0], lineStartPoint[1]), 2)  # acceleration
            if endPoint[2] < 0:
                currentPlanet.rotateModel(totalxrotation, totalyrotation)
                currentPlanet.draw()  # redraw planet if it is in front of force arrow

        # draw marker
        pygame.draw.circle(screen, WHITE, (planetPosition[0], planetPosition[1]), 20, 3)
        pygame.draw.lines(screen, WHITE, False, ((planetPosition[0], planetPosition[1] + 20),
                                                 (planetPosition[0], planetPosition[1] + 30),
                                                 (planetPosition[0] + 15, planetPosition[1] + 40)),
                          3)  # arrow
        screen.blit(currentPlanet.smallRenderedName, (planetPosition[0] + 16, planetPosition[1] + 40))


    # drawing menu
    screen.blit(menuImage, (0, 0))
    screen.blit(title, (6, 14))
    pygame.draw.rect(screen, SPACEYELLOW, menuborder, width=borderwidth, border_radius=2 * borderwidth)
    pygame.draw.rect(screen, LIGHTGREY, factBox, border_radius=10)
    screen.blit(factText, (16, 370))

    if currentPlanet is not None:# menu side - must be drawn after menu boxes
        # display txt
        currentPlanet.visitButton.draw()
        currentPlanet.display_facts()
        screen.blit(currentPlanet.renderedName, (16, 215))
        screen.blit(visitText1, (180, 303))
        screen.blit(visitText2, (180, 317))
        screen.blit(font.render('This planet has completed', True, WHITE), (16, 250))
        screen.blit(font.render(str(currentPlanet.orbitCount) + ' full orbits around the Sun', True, WHITE), (16, 263))

        # edit/update slider values
        currentPlanet.radiusSlider.calcValue()
        currentPlanet.densitySlider.calcValue()
        currentPlanet.set_radius(currentPlanet.radiusSlider.value)
        currentPlanet.set_density(currentPlanet.densitySlider.value)

        # additional features for custom planets
        if currentPlanet.custom:
            currentPlanet.redSlider.draw()
            currentPlanet.greenSlider.draw()
            currentPlanet.blueSlider.draw()
            currentPlanet.deleteButton.draw()
            screen.blit(deleteText, (170, 618))
            currentPlanet.pcolour = (int(abs(currentPlanet.redSlider.value)), int(abs(currentPlanet.greenSlider.value)), int(abs(currentPlanet.blueSlider.value)))
            pygame.draw.circle(screen, currentPlanet.pcolour, (125, 560), 2 * currentPlanet.visibleradius)
            if isinstance(currentPlanet, RingedPlanet):
                pygame.draw.line(screen, currentPlanet.scolour, (125 - 3 * currentPlanet.visibleradius, 560), (125 + 3 * currentPlanet.visibleradius, 560), 7)
            if currentPlanet.deleteButton.state:
                planets.remove(currentPlanet)
                planetsToDraw.remove(currentPlanet)
                selectablePlanets.remove(currentPlanet)
                del currentPlanet
                planetNumber -= 1
                currentPlanet = selectablePlanets[planetNumber]
                if currentPlanet.custom:
                    sliders = [currentPlanet.radiusSlider, currentPlanet.densitySlider,
                               currentPlanet.redSlider, currentPlanet.greenSlider, currentPlanet.blueSlider]
                else:
                    sliders = [currentPlanet.radiusSlider, currentPlanet.densitySlider]
    else:# if there is no planet selected, display facts about the whole solar system
        row = 0
        for fact in SolarSystemFacts:
            screen.blit(fact, (16, 400 + row * 15))
            row += 1


    for slider in sliders:
        if slider.sliding:
            slider.move(mouseChange)
        slider.draw()

    # buttons and labels
    for button in cycleButtons:
        button.draw()
    pygame.draw.polygon(screen, SPACEYELLOW, ((80, 165), (110, 150), (110, 180)))
    pygame.draw.polygon(screen, SPACEYELLOW, ((145, 150), (175, 165), (145, 180)))
    plotButton.draw()
    screen.blit(plotText, (24, 68))
    createButton.draw()
    screen.blit(createText1, (108, 63))
    screen.blit(createText2, (99, 75))
    clearButton.draw()
    screen.blit(clearText1, (181, 63))
    screen.blit(clearText2, (173, 75))
    excelButton.draw()
    screen.blit(excelText, (24, 108))
    hideButton.draw()
    screen.blit(hideText1, (108, 103))
    screen.blit(hideText2, (96, 115))
    viewButton.draw()
    screen.blit(viewText1, (170, 103))
    screen.blit(viewText2, (181, 115))
    # menu and border
    screen.blit(font.render(str(len(meteors)) + ' meteors', True, SPACEYELLOW), (16, 660))
    screen.blit(bottomText1, (16, 672))
    screen.blit(bottomText2, (16, 690))
    screen.blit(bottomText3, (16, 708))
    screen.blit(bottomText4, (16, 726))

    screen.blit(font.render((str(round((timeElapsed / (60*60*24*365.25)), 2)) + ' years since start'), True, WHITE), (menuwidth + 10, 10))
    screen.blit(font.render(str(round(clock.get_fps(), 1)) + ' FPS', True, WHITE), (940, 10))
    pygame.display.update()

pygame.quit()
print('Time Elapsed:')
print(str(timeElapsed / (60*60*24*365.25)) + 'years')
